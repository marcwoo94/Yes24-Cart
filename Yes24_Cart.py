import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Color
from bs4 import BeautifulSoup
from selenium import webdriver
import os


"""------------------------------------ Log In to Yes24 --------------------------------------"""


driver = webdriver.Chrome("C:\\Users\\user\\PycharmProjects\\untitled\\chromedriver.exe")
driver.implicitly_wait(3)
driver.get("https://www.yes24.com/Templates/FTLogin.aspx")
driver.maximize_window()
driver.find_element_by_name("SMemberID").send_keys("dnwpgus")
driver.find_element_by_name("SMemberPassword").send_keys("424-3109")
driver.find_element_by_xpath("//*[@id='btnLogin']/span/em").click()


"""------------------------------------ Extract Link List ------------------------------------"""


def get_product_info(book):
    link = book.find("a", class_="pd_a")["href"]
    return link

def get_page_info():
    driver.get("http://ssl.yes24.com/Cart/Cart")
    result = driver.page_source
    bs_obj = BeautifulSoup(result, "html.parser")
    books = bs_obj.find_all("td", class_="le")
    link_list = [get_product_info(book) for book in books]
    return link_list

result = []

def get_links():
    result.extend(get_page_info())
    return result


"""---------------------------------Extract Book Information---------------------------------"""


def get_book_info(url):
    result = requests.get(url)
    bs_obj = BeautifulSoup(result.content, "html.parser")

    title = bs_obj.find("h2", class_="gd_name").text  # 제목
    gd_pub = bs_obj.find("span", class_="gd_pub")  # 출판사
    publisher = gd_pub.find("a").text
    table_price = bs_obj.find("div", class_="gd_infoTb")  # 가격
    price = table_price.find_all("em", class_="yes_m")[0].text.split("원")[0]
    isbn = bs_obj.find_all("td", class_="txt lastCol")[2].text  # ISBN
    date = bs_obj.find_all("td", class_="txt lastCol")[0].text  # 출판일
    contributors = bs_obj.find("span", class_="gd_auth")
    try: author = contributors.find_all("a")[0].text  # 저자
    except:
        try: author = contributors.text.strip()
        except: author = ""
    try: translator = contributors.find_all("a")[1].text  # 역자
    except: translator = ""
    try: cover = bs_obj.find("span", class_="gd_feature").text.split("[")[1].split("]")[0].strip()  # 표지
    except: cover = ""
    category_table = bs_obj.find("dl", class_="yesAlertDl")
    category_list = category_table.find_all("li")
    try: category = category_list[0].find_all("a")[1].text + ", " + category_list[1].find_all("a")[1].text + ", " + category_list[2].find_all("a")[1].text
    except:
        try: category = category_list[0].find_all("a")[1].text + ", " + category_list[1].find_all("a")[1].text
        except: category = category_list[0].find_all("a")[1].text

    dictionary1 = {}
    dictionary1["title"] = title
    dictionary1["author"] = author
    dictionary1["translator"] = translator
    dictionary1["publisher"] = publisher
    dictionary1["price"] = price
    dictionary1["cover"] = cover
    dictionary1["ISBN"] = isbn
    dictionary1["date"] = date
    dictionary1["category"] = category

    return dictionary1

urls = get_links()

def get_info():
    book_info = [get_book_info(url) for url in urls]
    return book_info


"""--------------------------------------Save in Excel--------------------------------------"""


def save_in_excel():
    j1 = get_info()
    j2 = json.dumps(j1)
    j3 = json.loads(j2)

    wb = Workbook()
    ws1 = wb.active
    ws1.freeze_panes = "A2"

    ws1["A1"] = "제목"
    ws1["A1"].font = Font(size=12, bold=True)
    ws1["A1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
    ws1["B1"] = "저자"
    ws1["B1"].font = Font(size=12, bold=True)
    ws1["B1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
    ws1["C1"] = "역자"
    ws1["C1"].font = Font(size=12, bold=True)
    ws1["C1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
    ws1["D1"] = "출판사"
    ws1["D1"].font = Font(size=12, bold=True)
    ws1["D1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
    ws1["E1"] = "가격"
    ws1["E1"].font = Font(size=12, bold=True)
    ws1["E1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
    ws1["F1"] = "표지"
    ws1["F1"].font = Font(size=12, bold=True)
    ws1["F1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
    ws1["G1"] = "ISBN"
    ws1["G1"].font = Font(size=12, bold=True)
    ws1["G1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
    ws1["H1"] = "출판일"
    ws1["H1"].font = Font(size=12, bold=True)
    ws1["H1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))

    for row_index in range(1, len(j1)+1):
        ws1.cell(row=row_index + 1, column=1).value = j3[row_index - 1]["title"]
        ws1.cell(row=row_index + 1, column=2).value = j3[row_index - 1]["author"]
        ws1.cell(row=row_index + 1, column=3).value = j3[row_index - 1]["translator"]
        ws1.cell(row=row_index + 1, column=4).value = j3[row_index - 1]["publisher"]
        ws1.cell(row=row_index + 1, column=5).value = j3[row_index - 1]["price"]
        ws1.cell(row=row_index + 1, column=6).value = j3[row_index - 1]["cover"]
        ws1.cell(row=row_index + 1, column=7).value = j3[row_index - 1]["ISBN"]
        ws1.cell(row=row_index + 1, column=8).value = j3[row_index - 1]["date"]

    wb.save(os.path.dirname(os.path.abspath(__file__)) + "\Yes24_Cart.xlsx")

save_in_excel()
